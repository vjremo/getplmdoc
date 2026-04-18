#!/usr/bin/env python3
"""Sync a CSP .properties file with a CSP RTM Excel file."""

import argparse
import os
import sys
from copy import copy

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)

HEADERS = ["FlextypePath", "Activity", "Action", "Client Side Plugin Type", "Client Side Plugin JSP File"]


def parse_properties(path):
    """Return list of (flextype, activity, action, csp_type, jsp_path) from active entries."""
    entries = []
    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip()
            parts = key.split(".")
            # Last part is the clientSidePluginType
            csp_type = parts[-1]
            # Second-to-last is Action
            action = parts[-2]
            # Third-to-last is Activity
            activity = parts[-3]
            # Everything before that is the FlextypePath
            flextype = ".".join(parts[:-3])
            entries.append((flextype, activity, action, csp_type, value))
    return entries


def read_rtm(path):
    """Return (workbook, worksheet, set of existing keys)."""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        return wb, ws, set()
    wb = load_workbook(path)
    ws = wb.active
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            existing.add((row[0], row[1], row[2], row[3]))
    return wb, ws, existing


def get_rtm_keys_with_rows(ws):
    """Return dict of (flextype, activity, action, csp_type) -> row_index for data rows."""
    result = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:
            result[(row[0], row[1], row[2], row[3])] = idx
    return result


def copy_row_style(ws, src_row, dest_row):
    for col in range(1, 6):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dest_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.alignment = copy(src.alignment)
            dst.border = copy(src.border)
            dst.number_format = src.number_format


def main():
    parser = argparse.ArgumentParser(description="Sync CSP .properties -> RTM xlsx")
    parser.add_argument("--properties", required=True, help="Path to .properties file")
    parser.add_argument("--rtm", required=True, help="Path to RTM .xlsx file")
    args = parser.parse_args()

    props = parse_properties(args.properties)
    wb, ws, existing_keys = read_rtm(args.rtm)
    rtm_keys = get_rtm_keys_with_rows(ws)

    added = []
    already_existed = []

    last_data_row = ws.max_row if ws.max_row >= 2 else None

    for flextype, activity, action, csp_type, jsp_path in props:
        key = (flextype, activity, action, csp_type)
        if key in existing_keys:
            already_existed.append(key)
        else:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=flextype)
            ws.cell(row=new_row, column=2, value=activity)
            ws.cell(row=new_row, column=3, value=action)
            ws.cell(row=new_row, column=4, value=csp_type)
            ws.cell(row=new_row, column=5, value=jsp_path)
            if last_data_row and last_data_row >= 2:
                copy_row_style(ws, last_data_row, new_row)
            last_data_row = new_row
            added.append(key)
            existing_keys.add(key)

    props_keys = {(f, a, ac, c) for f, a, ac, c, _ in props}
    orphans = [k for k in rtm_keys if k not in props_keys]

    wb.save(args.rtm)

    print(f"\nADDED ({len(added)}):")
    for k in added:
        print(f"  {k[0]}.{k[1]}.{k[2]}.{k[3]}")

    print(f"\nALREADY EXISTED ({len(already_existed)}):")
    for k in already_existed:
        print(f"  {k[0]}.{k[1]}.{k[2]}.{k[3]}")

    print(f"\nORPHANS in RTM (no matching active properties entry) ({len(orphans)}):")
    for k in orphans:
        print(f"  {k[0]}.{k[1]}.{k[2]}.{k[3]}")

    print(f"\nDone. RTM saved to: {args.rtm}")


if __name__ == "__main__":
    main()
