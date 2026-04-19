#!/usr/bin/env python3
"""Combined RTM runner: SSP, JSP, CSP, and Techpack."""

import argparse
import subprocess
import sys
from pathlib import Path

BASE_DIR = Path(__file__).parent
SCRIPTS = {
    "ssp": BASE_DIR / "ssp-rtm-sync" / "ssp.py",
    "jsp": BASE_DIR / "jsp-rtm-sync" / "scripts" / "jsp.py",
    "csp": BASE_DIR / "csp-rtm-sync" / "scripts" / "csp.py",
    "techpack": BASE_DIR / "techpack-rtm-sync" / "techpack.py",
}

# Sheet name used in the combined workbook for each module
SHEET_NAMES = {
    "ssp":      "SSP",
    "jsp":      "JSP",
    "csp":      "CSP",
    "techpack": "Techpack",
}


def section(title):
    print(f"\n{'='*60}", flush=True)
    print(f"  {title}", flush=True)
    print(f"{'='*60}", flush=True)


SKIPPED = None  # sentinel returned by runners when input files are absent


def run(cmd, cwd=BASE_DIR):
    result = subprocess.run(cmd, cwd=cwd)
    if result.returncode != 0:
        print(f"[ERROR] Command exited with code {result.returncode}", file=sys.stderr)
    return result.returncode


def _skip(label: str, missing: list[Path]) -> None:
    for f in missing:
        print(f"  [SKIP] Required file not found: {f}", flush=True)
    print(f"  [SKIP] {label} — skipped.", flush=True)
    return SKIPPED


def run_ssp(args):
    section("SSP")
    if args.ssp_input:
        missing = [Path(f) for f in args.ssp_input if not Path(f).exists()]
        if missing:
            return _skip("SSP", missing)
    else:
        found = sorted((BASE_DIR / "properties").glob("*.properties"))
        if not found:
            return _skip("SSP", [BASE_DIR / "properties" / "*.properties"])
    cmd = [sys.executable, str(SCRIPTS["ssp"]), "-o", args.ssp_output]
    if args.ssp_input:
        cmd += args.ssp_input
    if args.ssp_template:
        cmd += ["-t", args.ssp_template]
    return run(cmd)


def run_jsp(args):
    section("JSP")
    props = args.work_dir / "properties"
    required = [
        props / "custom.urlMappings.properties",
        props / "custom.activityControllerMappings.properties",
        props / "custom.controllerAliases.properties",
    ]
    present = [f for f in required if f.exists()]
    if not present:
        return _skip("JSP", required)
    return run([sys.executable, str(SCRIPTS["jsp"])], cwd=args.work_dir)


def run_csp(args):
    section("CSP")
    props_file = Path(args.csp_properties)
    if not props_file.exists():
        return _skip("CSP", [props_file])
    cmd = [
        sys.executable, str(SCRIPTS["csp"]),
        "--properties", args.csp_properties,
        "--rtm", args.csp_rtm,
    ]
    return run(cmd)


def run_techpack(args):
    section("Techpack")
    props = args.work_dir / "properties"
    required = [
        props / "ProductSpecification2.properties",
        props / "ProductSpecificationBOM2.properties",
		props / "ProductSpecificationMeasure2.properties",
    ]
    missing = [f for f in required if not f.exists()]
    if missing:
        return _skip("Techpack", missing)
    return run([sys.executable, str(SCRIPTS["techpack"])], cwd=args.work_dir)


def resolve_output_paths(args) -> dict[str, Path]:
    """Return the expected output file path for each module."""
    return {
        "ssp":      (BASE_DIR / args.ssp_output).resolve(),
        "jsp":      (args.work_dir / "JSP.xlsx").resolve(),
        "csp":      (BASE_DIR / args.csp_rtm).resolve(),
        "techpack": (args.work_dir / "Techpack.xlsx").resolve(),
    }


def _apply_uniform_style(ws) -> None:
    """Apply consistent blue-header / alternating-row style to every cell in a worksheet."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", start_color="4472C4")
    even_fill   = PatternFill("solid", start_color="DCE6F1")
    odd_fill    = PatternFill("solid", start_color="FFFFFF")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    data_font   = Font(name="Arial", size=11)
    h_center    = Alignment(horizontal="center", vertical="center")
    v_center    = Alignment(vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = h_center
            else:
                cell.font      = data_font
                cell.fill      = even_fill if cell.row % 2 == 0 else odd_fill
                cell.alignment = v_center


def merge_rtm(modules_run: list[str], output_paths: dict[str, Path], combined_path: Path) -> None:
    """Copy each module's RTM xlsx into one combined workbook with uniform styling."""
    try:
        import openpyxl
    except ImportError:
        print("[SKIP] openpyxl not installed — combined RTM not generated.", file=sys.stderr)
        return

    section("Combined RTM")

    combined = openpyxl.Workbook()
    combined.remove(combined.active)  # drop the default empty sheet

    sheets_added = 0
    for module in modules_run:
        src_path = output_paths[module]
        sheet_name = SHEET_NAMES[module]

        if not src_path.exists():
            print(f"  [SKIP] {src_path.name} not found — '{sheet_name}' omitted")
            continue

        src_wb = openpyxl.load_workbook(src_path)
        src_ws = src_wb.active
        dest_ws = combined.create_sheet(title=sheet_name)

        # Column widths and row heights
        for col_letter, dim in src_ws.column_dimensions.items():
            dest_ws.column_dimensions[col_letter].width = dim.width
        for row_idx, dim in src_ws.row_dimensions.items():
            dest_ws.row_dimensions[row_idx].height = dim.height

        # Copy values only (styles applied uniformly below)
        for row in src_ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)

        # Apply consistent style across all sheets
        _apply_uniform_style(dest_ws)

        data_rows = sum(
            1 for row in src_ws.iter_rows(min_row=2)
            if any(c.value is not None for c in row)
        )
        print(f"  [{sheet_name}] {data_rows} data rows copied from {src_path.name}")
        sheets_added += 1

    if sheets_added == 0:
        print("  No source files found — combined RTM not saved.")
        return

    combined.save(combined_path)
    print(f"\n  Saved {sheets_added} sheet(s) -> {combined_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Run all RTM update scripts: SSP, JSP, CSP, and Techpack."
    )

    parser.add_argument(
        "--only", nargs="+", choices=["ssp", "jsp", "csp", "techpack"],
        metavar="SCRIPT",
        help="Run only the specified scripts (default: all). Choices: ssp, jsp, csp, techpack",
    )
    parser.add_argument(
        "--work-dir", default=str(BASE_DIR),
        help="Working directory for JSP and Techpack scripts (default: project root)",
    )
    parser.add_argument(
        "--combined-output", default="CustomizationReport.xlsx",
        metavar="FILE", help="Output path for the combined RTM workbook (default: CustomizationReport.xlsx)",
    )
    parser.add_argument(
        "--no-combine", action="store_true",
        help="Skip generating the combined RTM workbook",
    )

    # SSP options
    ssp_group = parser.add_argument_group("SSP (ssp-rtm-sync/ssp.py)")
    ssp_group.add_argument("--ssp-input", nargs="*", default=[],
                           metavar="FILE",
                           help="SSP .properties input file(s). Defaults to all *.properties in properties/ folder.")
    ssp_group.add_argument("--ssp-output", default="SSP.xlsx",
                           metavar="FILE", help="SSP output .xlsx file (default: SSP.xlsx)")
    ssp_group.add_argument("--ssp-template", default=None,
                           metavar="FILE", help="Optional existing SSP.xlsx to use as template")

    # CSP options
    csp_group = parser.add_argument_group("CSP (csp-rtm-sync/scripts/csp.py)")
    csp_group.add_argument(
        "--csp-properties",
        default="properties/custom.clientSidePluginManagerMappings.properties",
        metavar="FILE", help="CSP .properties input file",
    )
    csp_group.add_argument("--csp-rtm", default="CSP.xlsx",
                           metavar="FILE", help="CSP .xlsx file (default: CSP.xlsx)")

    args = parser.parse_args()
    args.work_dir = Path(args.work_dir).resolve()

    scripts_to_run = args.only or ["ssp", "jsp", "csp", "techpack"]
    runners = {"ssp": run_ssp, "jsp": run_jsp, "csp": run_csp, "techpack": run_techpack}

    errors = 0
    skipped = []
    ran = []
    for name in scripts_to_run:
        result = runners[name](args)
        if result is SKIPPED:
            skipped.append(name)
        elif result != 0:
            errors += 1
        else:
            ran.append(name)

    print(f"\n{'='*60}", flush=True)
    if skipped:
        print(f"  Skipped : {', '.join(skipped)}")
    if errors:
        print(f"  Done with {errors} error(s).")
        sys.exit(1)
    else:
        print(f"  Completed: {', '.join(ran) if ran else 'none'}")

    if not args.no_combine and ran:
        output_paths = resolve_output_paths(args)
        combined_path = (BASE_DIR / args.combined_output).resolve()
        merge_rtm(ran, output_paths, combined_path)


if __name__ == "__main__":
    main()
