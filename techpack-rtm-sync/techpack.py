import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RTM_FILE = 'Techpack.xlsx'
PROPS_DIR = 'properties'

def parse_properties(filepath):
    entries = []
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('//'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                key, val = key.strip(), val.strip()
                if 'com.lcs' not in val:
                    entries.append((key, val))
    return entries

entries = parse_properties(f'{PROPS_DIR}/ProductSpecification2.properties') + \
          parse_properties(f'{PROPS_DIR}/ProductSpecificationBOM2.properties')

if os.path.exists(RTM_FILE):
    wb = openpyxl.load_workbook(RTM_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Techpack RTM'
    ws.append(['Property Key', 'Class'])

thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

for cell in ws[1]:
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='4472C4')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for i, (key, val) in enumerate(entries, start=2):
    fill_color = 'DCE6F1' if i % 2 == 0 else 'FFFFFF'
    for col, v in enumerate([key, val], start=1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.font = Font(name='Arial', size=11)
        cell.fill = PatternFill('solid', start_color=fill_color)
        cell.alignment = Alignment(vertical='center')
        cell.border = border

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 65

wb.save(RTM_FILE)
print(f"Updated Techpack.xlsx with {len(entries)} entries.")
for k, v in entries:
    print(f"  {k} => {v}")
