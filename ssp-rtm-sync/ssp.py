#!/usr/bin/env python3
"""Convert custom.lcs.plugins.properties to an RTM Excel file."""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


HEADERS = [
    "Plugin Number",
    "Target Class",
    "Target Type",
    "Plugin Class",
    "Plugin Method",
    "Event",
    "Priority",
    "Bypass Security",
]

COL_WIDTHS = [12.5, 22.0, 10.0, 42.0, 24.0, 22.0, 8.0, 16.0]


def parse_properties(path: Path) -> list[dict]:
    rows = []
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            key, _, value = line.partition("=")
            if not key.startswith("com.lcs.wc.foundation.LCSPluginManager"):
                continue
            plugin_num = int(key.split(".")[-1])
            parts = dict(p.split("|", 1) for p in value.split("^"))
            target_class_fq = parts.get("targetClass", "")
            bypass_raw = parts.get("bypassSecurity", "").strip().lower()
            bypass = bypass_raw == "true" if bypass_raw else False
            rows.append({
                "plugin_num": plugin_num,
                "target_class": target_class_fq.split(".")[-1],
                "target_type": parts.get("targetType", ""),
                "plugin_class": parts.get("pluginClass", ""),
                "plugin_method": parts.get("pluginMethod", ""),
                "event": parts.get("event", ""),
                "priority": int(parts.get("priority", "1")),
                "bypass_security": bypass,
            })
    return sorted(rows, key=lambda r: (r["target_class"], r["priority"], r["plugin_num"]))


def write_rtm(rows: list[dict], out_path: Path, template: Path | None = None) -> None:
    if template and template.exists():
        wb = openpyxl.load_workbook(template)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SSP RTM"
        ws.append(HEADERS)
        for i, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r["plugin_num"])
        ws.cell(i, 2, r["target_class"])
        ws.cell(i, 3, r["target_type"])
        ws.cell(i, 4, r["plugin_class"])
        ws.cell(i, 5, r["plugin_method"])
        ws.cell(i, 6, r["event"])
        ws.cell(i, 7, r["priority"])
        ws.cell(i, 8, r["bypass_security"])

    wb.save(out_path)
    print(f"Wrote {len(rows)} rows -> {out_path}")


def main():
    parser = argparse.ArgumentParser(description="LCS plugins.properties → SSP_RTM.xlsx")
    parser.add_argument("input", nargs="*", default=[],
                        help="Path(s) to .properties file(s). Defaults to all *.properties in cwd.")
    parser.add_argument("-o", "--output", default="SSP_RTM.xlsx", help="Output .xlsx path")
    parser.add_argument("-t", "--template", default=None,
                        help="Optional existing SSP_RTM.xlsx to use as template")
    args = parser.parse_args()

    inputs = args.input or sorted((Path(".") / "properties").glob("*.properties"))
    if not inputs:
        sys.exit("No .properties files found.")

    rows = []
    for inp in inputs:
        props_path = Path(inp)
        if not props_path.exists():
            sys.exit(f"File not found: {props_path}")
        found = parse_properties(props_path)
        if found:
            print(f"  {props_path}: {len(found)} LCS plugin entries")
        rows.extend(found)

    if not rows:
        sys.exit("No com.lcs.wc.foundation.LCSPluginManager entries found in properties file(s).")

    template = Path(args.template) if args.template else None
    write_rtm(rows, Path(args.output), template)


if __name__ == "__main__":
    main()
