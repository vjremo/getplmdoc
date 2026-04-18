import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_properties(filepath):
    entries = []
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('//') or line.startswith('#'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                entries.append((key.strip(), val.strip()))
    return entries

def write_rtm(ws, entries, col_a_width=45):
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for cell in ws[1]:
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, (key, val) in enumerate(entries, start=2):
        fill_color = 'DCE6F1' if i % 2 == 0 else 'FFFFFF'
        for col, v in enumerate([key, val], start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = Font(name='Arial', size=11)
            cell.fill = PatternFill('solid', start_color=fill_color)
            cell.alignment = Alignment(vertical='center')
            cell.border = border

    ws.column_dimensions['A'].width = col_a_width
    ws.column_dimensions['B'].width = 65

# --- URL Mappings ---
url_entries = parse_properties('custom.urlMappings.properties')

# --- Activity Controller Mappings (resolved through aliases) ---
activity_map = dict(parse_properties('custom.activityControllerMappings.properties'))
alias_map = dict(parse_properties('custom.controllerAliases.properties'))
controller_entries = [
    (f"{activity_key},{controller_alias}", alias_map.get(controller_alias, ''))
    for activity_key, controller_alias in activity_map.items()
]

all_entries = url_entries + controller_entries

wb = openpyxl.load_workbook('JSP_RTM.xlsx')
write_rtm(wb.active, all_entries, col_a_width=55)
wb.save('JSP_RTM.xlsx')
print(f"Updated JSP_RTM.xlsx with {len(all_entries)} entries ({len(url_entries)} URL mappings + {len(controller_entries)} controller mappings).")
