#!/usr/bin/env python3
"""
CSP RTM Sync Script
Syncs the CSP RTM (Requirements Traceability Matrix) with the properties file.

Usage:
    python sync_csp_rtm.py --properties <path> --rtm <path> [--output <path>]
"""

import argparse
import sys
import shutil
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)


def parse_properties_file(properties_path):
    """Parse the .properties file and return list of active (non-commented) CSP entries."""
    entries = []
    with open(properties_path, 'r') as f:
        for line_num, line in enumerate(f, 1):
            stripped = line.strip()
            # Skip blank lines and comments (lines starting with # or ##)
            if not stripped or stripped.startswith('#'):
                continue
            # Parse key=value
            if '=' in stripped:
                key, _, value = stripped.partition('=')
                key = key.strip()
                value = value.strip()
                entries.append({
                    'line': line_num,
                    'key': key,
                    'value': value,
                    'raw': stripped
                })
    return entries


def inspect_rtm(wb):
    """Inspect the RTM workbook and return info about sheets and columns."""
    info = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Read header row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        info[sheet_name] = {
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'headers': headers
        }
    return info


def find_key_column(headers, candidates=None):
    """Find the column index (0-based) that likely contains CSP keys/entries."""
    if candidates is None:
        candidates = ['key', 'property', 'entry', 'csp', 'mapping', 'path', 'flextype', 'name']
    for i, h in enumerate(headers):
        if h and any(c in str(h).lower() for c in candidates):
            return i
    return None


def sync_rtm_with_properties(rtm_path, properties_path, output_path=None):
    """
    Main sync function.
    Returns a dict with keys: added, already_existed, orphans, summary
    """
    results = {
        'added': [],
        'already_existed': [],
        'orphans': [],
        'summary': ''
    }

    # Parse properties file
    prop_entries = parse_properties_file(properties_path)
    prop_keys = {e['key']: e for e in prop_entries}

    print(f"Properties file: {len(prop_entries)} active entries found")
    for e in prop_entries:
        print(f"  [{e['line']}] {e['key']} = {e['value']}")

    # Try to open the RTM
    try:
        wb = load_workbook(rtm_path)
    except Exception as ex:
        print(f"WARNING: Could not open {rtm_path}: {ex}")
        print("Trying to work on a copy at /tmp/CSP_RTM_eval1_base.xlsx ...")
        copy_path = '/tmp/CSP_RTM_eval1_base.xlsx'
        shutil.copy2(rtm_path, copy_path)
        wb = load_workbook(copy_path)
        rtm_path = copy_path

    # Inspect RTM
    rtm_info = inspect_rtm(wb)
    print(f"\nRTM Sheets: {list(rtm_info.keys())}")
    for sheet, info in rtm_info.items():
        print(f"  Sheet '{sheet}': {info['max_row']} rows, headers={info['headers']}")

    # Use the first sheet
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    print(f"\nWorking with sheet: '{sheet_name}'")
    print(f"Headers: {headers}")

    # Find the key/property column
    key_col_idx = find_key_column(headers)
    if key_col_idx is None:
        # Default: try first column
        key_col_idx = 0
        print(f"No obvious key column found, defaulting to column index 0 ('{headers[0]}')")
    else:
        print(f"Key column: index {key_col_idx} ('{headers[key_col_idx]}')")

    # Also find value/JSP column
    value_col_candidates = ['value', 'jsp', 'file', 'path', 'plugin']
    value_col_idx = find_key_column(headers, value_col_candidates)
    if value_col_idx is None:
        value_col_idx = 1 if len(headers) > 1 else 0
        print(f"No obvious value column found, defaulting to column index {value_col_idx}")
    else:
        print(f"Value column: index {value_col_idx} ('{headers[value_col_idx]}')")

    # Collect existing RTM keys (rows 2+)
    rtm_keys = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[key_col_idx] is not None:
            rtm_keys[str(row[key_col_idx]).strip()] = row_idx

    print(f"\nExisting RTM entries: {len(rtm_keys)}")
    for k, r in rtm_keys.items():
        print(f"  Row {r}: '{k}'")

    # Sync: check each property entry against RTM
    for key, entry in prop_keys.items():
        if key in rtm_keys:
            results['already_existed'].append({
                'key': key,
                'value': entry['value'],
                'rtm_row': rtm_keys[key]
            })
        else:
            # Need to add this entry to the RTM
            new_row = ws.max_row + 1
            # Build new row data with empty cells for all columns
            row_data = [None] * len(headers)
            row_data[key_col_idx] = key
            if value_col_idx < len(row_data):
                row_data[value_col_idx] = entry['value']
            ws.append(row_data)
            results['added'].append({
                'key': key,
                'value': entry['value'],
                'rtm_row': new_row
            })

    # Find orphans: RTM entries not in properties
    for rtm_key in rtm_keys:
        if rtm_key not in prop_keys:
            results['orphans'].append({
                'key': rtm_key,
                'rtm_row': rtm_keys[rtm_key]
            })

    # Save RTM
    if output_path:
        save_path = output_path
    else:
        save_path = rtm_path
    wb.save(save_path)
    print(f"\nRTM saved to: {save_path}")

    # Build summary
    lines = []
    lines.append("=" * 70)
    lines.append("CSP RTM SYNC RESULTS")
    lines.append("=" * 70)
    lines.append(f"Properties file: {properties_path}")
    lines.append(f"RTM file: {save_path}")
    lines.append("")

    lines.append(f"ADDED ({len(results['added'])} entries):")
    if results['added']:
        for item in results['added']:
            lines.append(f"  + Row {item['rtm_row']}: {item['key']} = {item['value']}")
    else:
        lines.append("  (none)")

    lines.append("")
    lines.append(f"ALREADY EXISTED ({len(results['already_existed'])} entries):")
    if results['already_existed']:
        for item in results['already_existed']:
            lines.append(f"  = Row {item['rtm_row']}: {item['key']}")
    else:
        lines.append("  (none)")

    lines.append("")
    lines.append(f"ORPHANS ({len(results['orphans'])} entries in RTM but not in properties):")
    if results['orphans']:
        for item in results['orphans']:
            lines.append(f"  ? Row {item['rtm_row']}: {item['key']}")
    else:
        lines.append("  (none)")

    lines.append("")
    lines.append("=" * 70)
    results['summary'] = '\n'.join(lines)
    return results


def main():
    parser = argparse.ArgumentParser(description='Sync CSP RTM with properties file')
    parser.add_argument('--properties', required=True, help='Path to .properties file')
    parser.add_argument('--rtm', required=True, help='Path to RTM .xlsx file')
    parser.add_argument('--output', help='Output path for updated RTM (default: overwrite RTM)')
    parser.add_argument('--result-txt', help='Path to save plain text result summary')
    args = parser.parse_args()

    print(f"Syncing RTM with properties file...")
    print(f"  Properties: {args.properties}")
    print(f"  RTM: {args.rtm}")

    results = sync_rtm_with_properties(args.rtm, args.properties, args.output)

    print("\n" + results['summary'])

    if args.result_txt:
        os.makedirs(os.path.dirname(args.result_txt), exist_ok=True)
        with open(args.result_txt, 'w') as f:
            f.write(results['summary'])
        print(f"\nResult summary saved to: {args.result_txt}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
