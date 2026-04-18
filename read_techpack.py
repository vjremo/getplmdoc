import openpyxl
wb = openpyxl.load_workbook('Techpack_RTM.xlsx')
print('Sheets:', wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n=== Sheet: {sname} (maxrow={ws.max_row}, maxcol={ws.max_column}) ===')
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True):
        print(row)
