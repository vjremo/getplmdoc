import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_properties(filepath):
    entries = {}
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('//') or line.startswith('#'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                entries[key.strip()] = val.strip()
    return entries

# Load both property files
activity_map = parse_properties('custom.activityControllerMappings.properties')
alias_map = parse_properties('custom.controllerAliases.properties')

# Resolve: Activity Key -> Controller Alias -> JSP File
entries = []
for activity_key, controller_alias in activity_map.items():
    jsp_file = alias_map.get(controller_alias, '')
    entries.append((f"{activity_key},{controller_alias}", jsp_file))

# Load existing workbook to preserve header style
wb = openpyxl.load_workbook('JSP_RTM.xlsx')
ws = wb.active

thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name='Arial', bold=True, color='FFFFFF')
header_fill = PatternFill('solid', start_color='4472C4')
data_font = Font(name='Arial', size=11)

# Clear data rows, preserve header
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Re-style header
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Write resolved entries
for i, (page_key, jsp_file) in enumerate(entries, start=2):
    fill_color = 'DCE6F1' if i % 2 == 0 else 'FFFFFF'
    for col, val in enumerate([page_key, jsp_file], start=1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = data_font
        cell.fill = PatternFill('solid', start_color=fill_color)
        cell.alignment = Alignment(vertical='center')
        cell.border = border

ws.column_dimensions['A'].width = 55
ws.column_dimensions['B'].width = 65

wb.save('JSP_RTM.xlsx')
print(f"Updated JSP_RTM.xlsx with {len(entries)} entries.")
for k, v in entries:
    print(f"  {k} => {v}")
