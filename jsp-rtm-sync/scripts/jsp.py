import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RTM_FILE = 'JSP_RTM.xlsx'
PROPS_DIR = 'properties'
URL_MAPPINGS = f'{PROPS_DIR}/custom.urlMappings.properties'
ACTIVITY_MAPPINGS = f'{PROPS_DIR}/custom.activityControllerMappings.properties'
CONTROLLER_ALIASES = f'{PROPS_DIR}/custom.controllerAliases.properties'

def parse_properties(filepath):
    entries = []
    if not os.path.exists(filepath):
        return entries
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('//') or line.startswith('#'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                entries.append((key.strip(), val.strip()))
    return entries

def write_rtm(ws, entries):
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for cell in ws[1]:
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, (key, val) in enumerate(entries, start=2):
        fill_color = 'DCE6F1' if i % 2 == 0 else 'FFFFFF'
        for col, v in enumerate([key, val], start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = Font(name='Arial', size=11)
            cell.fill = PatternFill('solid', start_color=fill_color)
            cell.alignment = Alignment(vertical='center')
            cell.border = border

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 65

def load_or_create_workbook(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['Page Key', 'JSP File'])
    return wb

url_entries = parse_properties(URL_MAPPINGS)

activity_map = dict(parse_properties(ACTIVITY_MAPPINGS))
alias_map = dict(parse_properties(CONTROLLER_ALIASES))
controller_entries = [
    (f"{activity_key},{controller_alias}", alias_map.get(controller_alias, ''))
    for activity_key, controller_alias in activity_map.items()
]

all_entries = url_entries + controller_entries

wb = load_or_create_workbook(RTM_FILE)
write_rtm(wb.active, all_entries)
wb.save(RTM_FILE)
print(f"Updated {RTM_FILE} with {len(all_entries)} entries "
      f"({len(url_entries)} URL mappings + {len(controller_entries)} controller mappings).")
